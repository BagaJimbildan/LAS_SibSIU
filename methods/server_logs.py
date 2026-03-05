from openpyxl import load_workbook
from datetime import datetime
import user_info as user_inf
import static_info as stat_inf
import file_master as file_m
import app_info as app_inf

path_log = None

def write_excel(action: str, note = None, is_local_log=False):
    try:
        if (path_log is not None and not is_local_log) or is_local_log:
            if is_local_log: wb = load_workbook(app_inf.local_log_path)
            else: wb = load_workbook(path_log)
            ws = wb['Лист1']

            last_row = ws.max_row
            free_row = last_row + 1

            if app_inf.write_server:
                data = [user_inf.ticket,
                        user_inf.username[1],
                        user_inf.current_room,
                        user_inf.room,
                        user_inf.owner,
                        user_inf.subdivision]
            else:
                data = [stat_inf.do_not_know]*6


            for col, value in enumerate(data+[str(datetime.now()), action, note if note is not None else stat_inf.default_note], start=1):
                ws.cell(row=free_row, column=col, value=value)

            if not is_local_log: wb.save(path_log)
            else: wb.save(file_m.excel_file)
            wb.close()

        else:
            return [2, "Сервер отключен или файл не найден" if not is_local_log else "Не найден файл для записи логов текущей сессии"]
        return [0,0]
    except Exception as e:
        return [2,e]

def test_open_excel():
    try:
        if path_log is not None:
            wb = load_workbook(path_log)
            wb.save(path_log)
            wb.close()
        else:
            return [1, "Сервер отключен или файл не найден"]
        return [0,0]
    except Exception as e:
        return [1,e]



