
from openpyxl import load_workbook

def write_to_first_free_row(filename, data):
    """
    Записывает список data (длиной 7) в первую полностью свободную строку
    на листе (столбцы A–G). Файл Excel не открывается на экране.
    """
    # Открываем книгу
    wb = load_workbook(filename)
    ws = wb.active  # или wb['Лист1'] для конкретного листа

    # Определяем максимальное количество строк с данными
    # (используем max_row, чтобы ограничить поиск)
    max_row = ws.max_row

    # Поиск первой строки, где все ячейки в столбцах 1–7 пустые
    free_row = None
    for row in range(1, max_row + 2):  # проверяем включая строку max_row+1
        all_empty = True
        for col in range(1, 8):        # столбцы 1..7 (A..G)
            if ws.cell(row, col).value is not None:
                all_empty = False
                break
        if all_empty:
            free_row = row
            break

    # Если не нашли пустой строки до max_row+1, используем max_row+1
    if free_row is None:
        free_row = max_row + 1

    # Записываем данные (предполагаем, что data — список из 7 элементов)
    for col, value in enumerate(data, start=1):
        ws.cell(row=free_row, column=col, value=value)

    # Сохраняем файл
    wb.save(filename)
    wb.close()
    print(f"Данные записаны в строку {free_row}.")

# Пример использования
if __name__ == "__main__":
    file = r"C:\Users\Baga\Desktop\тест сервер файл.xlsx"
    info = ["AZ-46", "гРИГОРЬЕКВ А.К.", "ПК-23", "12.02.2004", 45.6, "ещё текст", "последний"]
    write_to_first_free_row(file, info)